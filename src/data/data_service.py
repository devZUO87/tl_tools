import openpyxl
from openpyxl.workbook import Workbook
from collections import defaultdict
from src.function.measurement import Measurement

class DataService:
    def __init__(self):
        """
        初始化 DataService 实例。
        data: 存储导入的原始数据（列表）。
        grouped_data: 存储分组后的数据（字典，键为分组元组，值为数据列表）。
        """
        self.data = []
        self.grouped_data = defaultdict(list)

    def import_excel(self, file_path):
        """
        从指定的 Excel 文件导入数据。
        参数：
            file_path (str): Excel 文件的路径。
        返回：
            list: 读取的每一行数据组成的列表，每行是一个元组。
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        return self.data

    def export_excel(self, data, file_path, calculated=False):
        """
        将数据导出为 Excel 文件。
        参数：
            data (list): 要导出的数据，每行为一个元组或列表。
            file_path (str): 保存的 Excel 文件路径。
            calculated (bool): 是否导出包含计算结果的表头。
        返回：
            None
        """
        wb = Workbook()
        ws = wb.active
        # 根据是否已计算，写入不同的表头
        if calculated:
            ws.append([
                "文件名", "测站", "目标", "归零方向均值", "天顶角均值", "斜距(m)", "仪器高(m)", "目标高(m)",
                "测站温度(℃)", "目标温度(℃)", "测站气压(hPa)", "目标气压(hPa)", "气象改正(m)", "加乘常数改正(m)",
                "平距(m)", "高差(m)", "高差中数(m)", "往返不符值(mm)", "限差(mm)"
            ])
        else:
            ws.append([
                "文件名", "测站", "目标", "归零方向均值", "天顶角均值", "斜距", "仪器高(m)", "目标高(m)", "测站温度",
                "目标温度", "测站气压", "目标气压"
            ])
        # 写入数据内容
        for row in data:
            clean_row = [cell if cell is not None else '' for cell in row]
            ws.append(clean_row)
        # 设置单元格居中
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        # 设置列宽
        column_widths = {'A': 20, 'B': 10, 'C': 10, 'D': 16, 'E': 20, 'F': 20, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        wb.save(file_path)

    def group_data(self, data):
        """
        对原始数据按测站和目标进行分组。
        参数：
            data (list): 原始数据列表，每行为一个元组。
        返回：
            dict: 分组后的数据，键为 (测站, 目标) 的元组，值为该组的数据列表。
        """
        grouped_data = defaultdict(list)
        for value in data:
            group_key = (value[1], value[2])  # 以测站和目标为分组依据
            grouped_data[group_key].append(value)
        return grouped_data

    def sort_and_calculate(self, grouped_data):
        """
        对分组后的数据进行配对排序，并对每组数据进行测量计算。
        参数：
            grouped_data (dict): 分组后的数据，键为 (测站, 目标)，值为数据列表。
        返回：
            dict: 排序并计算后的分组数据，结构同输入。
        主要逻辑：
            - 先将有互为 (a, b) 和 (b, a) 的分组配对排序。
            - 对每组数据调用 Measurement.calculate_all 进行测量计算，
              并将结果追加到原始数据后面。
        """
        paired_data = []  # 存储有对应 (b, a) 键的键对
        unpaired_data = []  # 存储没有对应 (b, a) 键的键
        visited = set()  # 标记已处理的键
        # 遍历所有分组
        for key in grouped_data.keys():
            if key not in visited:
                reverse_key = (key[1], key[0])  # 生成 (b, a) 键
                if reverse_key in grouped_data:
                    paired_data.append((key, grouped_data[key]))
                    paired_data.append((reverse_key, grouped_data[reverse_key]))
                    visited.add(key)
                    visited.add(reverse_key)
                else:
                    unpaired_data.append((key, grouped_data[key]))
                    visited.add(key)
        # 合并有配对和无配对的数据
        sorted_data = paired_data + unpaired_data
        new_grouped_data = {k: v for k, v in sorted_data}
        # 对每组数据进行测量计算
        for key, data_list in new_grouped_data.items():
            updated_data = []
            for item in data_list:
                # 调用测量计算逻辑，返回字典
                results = Measurement.calculate_all(
                    s=float(item[5]), z=float(item[4]), i=float(item[6]), l=float(item[7]),
                    t_a=float(item[8]), t_b=float(item[9]), p_a=float(item[10]), p_b=float(item[11])
                )
                additional_values = tuple(results.values())  # 取所有计算结果
                updated_item = item + additional_values  # 拼接到原始数据后
                updated_data.append(updated_item)
            new_grouped_data[key] = updated_data
        return new_grouped_data