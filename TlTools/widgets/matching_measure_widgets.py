import openpyxl
from PyQt5.QtWidgets import QMainWindow, QMessageBox, QFileDialog
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from TlTools.widgets.menu_component import MenuComponent


class MatchingMeasureWidget(QMainWindow):
    def __init__(self, main_window):
        super().__init__()
        self.data = None
        self.main_window = main_window  # 传入主窗口实例
        self.import_button = None
        self.label = None
        self.menu_component = MenuComponent(self)
        self.setMenuBar(self.menu_component)
        self.import_data = None
        self.initUI()


    def initUI(self):
        self.setWindowTitle('导入外业手簿')
        self.setGeometry(400, 400, 300, 200)

        data_menu = self.menu_component.add_menu('数据')
        self.menu_component.add_action(data_menu, '导出匹配表', '导出匹配表', self.exportData)
        self.menu_component.add_action(data_menu, '导入外业手簿', '导入外业手簿', self.importData)


    def exportData(self):
        try:
            data = self.main_window.drag_table_widget.all_table_data
            wb = Workbook()
            ws = wb.active
            ws.append(
                ["文件名", "测站", "目标", "归零方向均值", "天顶角均值", "斜距" ,"仪器高(m)", "目标高(m)", "测站温度", "目标温度", "测站气压", "目标气压"])
            # 写入数据
            for row in data:
                # 处理 None 值，避免写入到 Excel
                clean_row = [cell if cell is not None else '' for cell in row]
                ws.append(clean_row)
            # 表格格式
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            # 宽度
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10

            # 保存为 Excel 文件
            wb.save('output.xlsx')
            QMessageBox.warning(self, '完成', '数据已保存到 Excel 文件')
        except Exception as e:
            QMessageBox.warning(self, '错误', e)

    def importData(self):
        print('导入外业手簿')
        path = QFileDialog.getOpenFileUrl(self, '选择文件')

        if path:
            # 调用主窗口的方法来创建并显示DraggableTableWidget
            self.data = self.get_excel_data(path)
            print(self.data)



    def get_excel_data(self,file_path):
        # Load the Excel workbook
        wb = load_workbook(file_path)
        ws = wb.active  # Get the active worksheet

        # Extract the data
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))  # Convert row to a list

        return data










