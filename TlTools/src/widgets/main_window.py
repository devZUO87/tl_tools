import math
import sys
from collections import defaultdict

import openpyxl
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QVBoxLayout,
    QWidget,
    QStatusBar,
    QMessageBox,
    QFileDialog,
    QTableWidgetItem,
    QTableWidget,
    QHeaderView,
)
from PyQt6.QtGui import QPalette, QColor
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook

from src.function.measurement import Measurement
from src.widgets.draggable_table_widgets import DraggableTableWidget
from src.widgets.import_data_widgets import ImportDataWindow
from src.widgets.menu_component import MenuComponent


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.matched = None # 匹配状态
        self.calculated = None  # 计算状态
        self.setWindowTitle("三角高程计算工具")
        self.init_ui()
        self.setup_styles()
        self.resize(800, 600)
           # 将窗口移动到屏幕中心
        screen = QApplication.primaryScreen().geometry()
        window_size = self.geometry()
        x = (screen.width() - window_size.width()) // 2
        y = (screen.height() - window_size.height()) // 2
        self.move(x, y)
        # 创建状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # 创建菜单栏
        self.menu_component = MenuComponent(self)
        self.setMenuBar(self.menu_component)  # 使用 setMenuBar 设置菜单
        self.initFileMenu()

        # 创建中心部件和布局
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # 添加导入数据窗口
        self.grouped_data = defaultdict(list)
        self.import_data_window = ImportDataWindow(self)
        # self.import_data_window.show()

        # Create a table widget
        self.table_widget = QTableWidget()
        self.data_keys = ["文件名", "测站", "目标", "归零方向均值", "天顶角均值", "斜距", "仪器高(m)", "目标高(m)",
                          "测站温度", "目标温度", "测站气压", "目标气压"]
        self.table_widget.setColumnCount(len(self.data_keys))
        self.table_widget.setHorizontalHeaderLabels(self.data_keys)
        self.layout.addWidget(self.table_widget)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.data = []


    def setup_styles(self):
            """设置全局样式"""
            self.setAutoFillBackground(True)
            palette = self.palette()
            palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 240))  # 浅灰色背景
            palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))    # 黑色文字
            palette.setColor(QPalette.ColorRole.Button, QColor(225, 225, 225))  # 按钮颜色
            palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))    # 按钮文字颜色
            self.setPalette(palette)

    def init_ui(self):
        # 现代化QSS样式
        self.setStyleSheet("""
        QWidget {
            background: #FFFFFF;
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            color: #222;
        }
        QGroupBox {
            border: 1px solid #E5E5E5;
            border-radius: 10px;
            margin-top: 16px;
            background: #fff;
        }
        QGroupBox:title {
            subcontrol-origin: margin;
            left: 12px;
            padding: 0 6px 0 6px;
            font-weight: bold;
            color: #0078D4;
            font-size: 11pt;
        }
        QPushButton {
            background-color: #0078D4;
            color: white;
            border: none;
            border-radius: 8px;
            padding: 10px 28px;
            font-size: 11pt;
            font-weight: 600;
        }
        QPushButton:hover {
            background-color: #005A9E;
        }
        QLineEdit, QComboBox {
            border: 1px solid #BDBDBD;
            border-radius: 8px;
            padding: 6px 10px;
            background: #fff;
            font-size: 10.5pt;
        }
        QComboBox::drop-down {
            subcontrol-origin: padding;
            subcontrol-position: top right;
            width: 28px;
            border: none;
            border-top-right-radius: 8px;
            border-bottom-right-radius: 8px;
            border-top-left-radius: 0px;
            border-bottom-left-radius: 0px;
            background: #E3F2FD;
            margin: 1px 1px 1px 0;
        }
        QComboBox QAbstractItemView {
            border: 1px solid #BDBDBD;
            background: #fff;
            selection-background-color: #E3F2FD;
            selection-color: #0078D4;
            border-radius: 8px;
            padding: 4px;
            margin: 0px;
            outline: none;        
        }
        QComboBox::view {
            margin: 0;
            padding: 0;
            border-radius: 8px;
            background: transparent;
        }               
        QLabel {
            font-size: 10.5pt;
        }
        QProgressBar {
            border: 1px solid #BDBDBD;
            border-radius: 8px;
            text-align: center;
            background: #E3F2FD;
            height: 18px;
        }
        QProgressBar::chunk {
            background-color: #0078D4;
            border-radius: 8px;
        }
        QCheckBox {
            spacing: 8px;
            font-size: 10pt;
        }
        """)

    def initFileMenu(self):
        file_menu = self.menu_component.add_menu('文件')
        self.menu_component.add_action(file_menu, '打开', '打开文件夹', self.openFile)
        self.menu_component.add_action(file_menu, '退出', '退出程序', self.close)

        import_menu = self.menu_component.add_menu('导入')
        self.menu_component.add_action(import_menu, '导入外业手簿', '导入', self.importMatchingTable)


        export_menu = self.menu_component.add_menu('导出')
        self.menu_component.add_action(export_menu, '导出外业测量数据', '导出外业测量数据', self.exportOutsideTable)
        self.menu_component.add_action(export_menu, '导出三角高程测量计算总表', '导出三角高程测量计算总表',
                                       self.exportOutsideTable)

        measure_menu = self.menu_component.add_menu('处理')
        self.menu_component.add_action(measure_menu, '匹配外业手簿', '匹配', self.set_draggable_table_widget)
        self.menu_component.add_action(measure_menu, '计算对象观测中误差', '对象观测计算', self.calculate_draggable_table_widget)


        help_menu = self.menu_component.add_menu('帮助')
        self.menu_component.add_action(help_menu, '关于', '关于程序', self.showAbout)

    def openFile(self):
        try:
            self.matched = False
            self.close()
            self.import_data_window.show()
            self.import_data_window.label.setText('请选择数据文件夹')
        except AttributeError:
            QMessageBox.warning(self, '错误', '没有可导入的数据')

    def importMatchingTable(self):
        try:
            file_dialog_result = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel Files (*.xlsx)')
            file_path = file_dialog_result[0]
            self.data = []
            if file_path:
                wb = load_workbook(file_path)
                ws = wb.active  # active 是属性，不是方法

                # 加载 Excel 数据
                for row in ws.iter_rows(min_row=2,values_only=True):
                    self.data.append(row)
                # 更新表格内容
                self.set_table_widget(self.data)
            else:
                QMessageBox.warning(self, '错误', '没有选择文件')
        except AttributeError:
            QMessageBox.warning(self, '错误', '没有可导入的数据')

    def set_table_widget(self, data):
        # 清空现有表格内容
        self.table_widget.clear()

        # 设置行数
        self.table_widget.setRowCount(len(data))

        # 设置列数
        self.table_widget.setColumnCount(len(self.data_keys))

        # 重新设置表头
        self.table_widget.setHorizontalHeaderLabels(self.data_keys)

        # 设置表格数据
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_data in enumerate(row_data):
                self.table_widget.setItem(row_idx, col_idx, QTableWidgetItem(str(cell_data)))

    def exportOutsideTable(self):
        # 导出匹配表
        if self.table_widget:
            try:
                data = self.get_all_table_data()
                self.save_data_to_excel(data)

            except Exception as e:
                msg_box = QMessageBox()
                msg_box.setIcon(QMessageBox.Icon.Warning)
                msg_box.setWindowTitle("警告")
                msg_box.setText(f"程序可能没有足够的权限写入文件。\n{str(e)}\n\n请以管理员身份运行程序。")
                msg_box.exec()
        else:
            QMessageBox.warning(self, '错误', '没有可导出的数据')

    def get_all_table_data(self):
        data = []
        for row in range(self.table_widget.rowCount()):
            if self.matched:
                row_data = [self.table_widget.item(row, col).text() if self.table_widget.item(row, col) else None for col in range(self.table_widget.columnCount())[1:-1]]
            else:
                row_data = [self.table_widget.item(row, col).text() if self.table_widget.item(row, col) else None for col in range(self.table_widget.columnCount())]
            data.append(row_data)
        return data

    def save_data_to_excel(self,data):
        wb = Workbook()
        ws = wb.active
        if self.calculated:
            ws.append(
                ["文件名", "测站", "目标", "归零方向均值", "天顶角均值", "斜距(m)", "仪器高(m)", "目标高(m)",
                 "测站温度(℃)", "目标温度(℃)", "测站气压(hPa)", "目标气压(hPa)", "气象改正(m)", "加乘常数改正(m)",
                 "平距(m)", "高差(m)", "高差中数(m)", "往返不符值(mm)", "限差(mm)"])
        else:
            ws.append(
                ["文件名", "测站", "目标", "归零方向均值", "天顶角均值", "斜距", "仪器高(m)", "目标高(m)", "测站温度",
                 "目标温度", "测站气压", "目标气压"])
        # 写入数据
        for row in data:
            # 处理 None 值，避免写入到 Excel
            clean_row = [cell if cell is not None else '' for cell in row]
            ws.append(clean_row)
        # 表格格式
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        # 宽度
        # 定义字典来存储各列的宽度
        column_widths = {'A': 20, 'B': 10, 'C': 10, 'D': 16, 'E': 20, 'F': 20, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12}

        # 使用循环设置列宽
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 保存为 Excel 文件
        file_dialog_result = QFileDialog.getSaveFileName(self, '选择文件夹和输入文件名', '', 'Excel Files (*.xlsx)')
        file_path = file_dialog_result[0]

        if file_path:
            # 保存文件
            wb.save(file_path)
            QMessageBox.warning(self, '完成', f'文件已保存到: {file_path}')

    def set_draggable_table_widget(self):
        if self.matched:
            QMessageBox.warning(self, '错误', '已匹配，无法重复匹配')
        else:
            self.grouped_data = defaultdict(list)
            self.matched = True
            self.get_grouped_data()
            self.sort_grouped_data()
            self.table_widget.clear()

            self.table_widget = DraggableTableWidget(self.grouped_data)
            central_widget = QWidget()
            layout = QVBoxLayout()
            layout.addWidget(self.table_widget)
            central_widget.setLayout(layout)

            # 将 central_widget 作为主窗口的中央组件
            self.setCentralWidget(central_widget)

    def get_grouped_data(self):
        # 根据 i1 和 i2 分组，返回分组后的数据
        for value in self.data:
            group_key = (value[1], value[2])
            self.grouped_data[group_key].append(value)

    def sort_grouped_data(self):
        paired_data = []  # 存储有对应 (b, a) 键的键对
        unpaired_data = []  # 存储没有对应 (b, a) 键的键
        visited = set()  # 用于标记已经处理过的键

        # 遍历 grouped_data
        for key in self.grouped_data.keys():
            if key not in visited:
                reverse_key = (key[1], key[0])  # 生成 (b, a) 键

                if reverse_key in self.grouped_data:
                    # 如果存在对应的 (b, a) 键，加入 paired_data
                    paired_data.append((key, self.grouped_data[key]))
                    paired_data.append((reverse_key, self.grouped_data[reverse_key]))
                    visited.add(key)
                    visited.add(reverse_key)
                else:
                    # 如果没有对应的 (b, a) 键，加入 unpaired_data
                    unpaired_data.append((key, self.grouped_data[key]))
                    visited.add(key)

        # 合并有对应 (b, a) 的数据和没有对应的数据
        sorted_data = paired_data + unpaired_data

        # 重新构建 grouped_data 按新顺序
        self.grouped_data = {k: v for k, v in sorted_data}

        # 遍历排序后的 grouped_data 并调用计算逻辑
        for key, data in self.grouped_data.items():
            updated_data = []  # 创建一个新列表来存储更新后的数据
            for item in data:
                # 生成 Measurement 实例并计算结果
                results = Measurement.calculate_all(
                    s=float(item[5]), z=float(item[4]), i=float(item[6]), l=float(item[7]),
                    t_a=float(item[8]), t_b=float(item[9]), p_a=float(item[10]), p_b=float(item[11])
                )

                # 将计算结果转换为元组
                additional_values = tuple(results.values())
                # print(additional_values)

                # 追加计算结果到原始项
                updated_item = item + additional_values

                # 将更新后的项添加到更新后的数据列表中
                updated_data.append(updated_item)

            # 更新 grouped_data 字典中的键
            self.grouped_data[key] = updated_data

            # 将计算的高度等中间结果更新到对应的 grouped_data 中
        # print(self.grouped_data)

    def calculate_draggable_table_widget(self):
        self.calculated = True
        if self.matched:
            for i in range(0, self.table_widget.rowCount(), 2):  # 每两行遍历
                if i + 1 >= self.table_widget.rowCount():  # 确保有下一行
                    continue

                # 获取当前行和下一行的第六列和第十四列数据
                def get_row_value(row, col):
                    item = self.table_widget.item(row, col)
                    return float(item.text()) if item else 0

                first_row_value = get_row_value(i, 16)
                second_row_value = get_row_value(i + 1, 16)
                first_row_value_1 = get_row_value(i, 14)
                second_row_value_1 = get_row_value(i + 1, 14)

                # 计算平均值、和与容差
                average_d =round( 0.5 * (first_row_value - second_row_value),5)
                sum_value = round((first_row_value + second_row_value) * 1000,5)
                tolerance = round(40 * math.sqrt(0.5 * (abs(first_row_value_1) + abs(second_row_value_1)) / 1000),5)

                # 将结果追加到当前行的指定列
                self.table_widget.setItem(i, self.table_widget.columnCount() - 4, QTableWidgetItem(str(average_d)))
                self.table_widget.setItem(i, self.table_widget.columnCount() - 3, QTableWidgetItem(str(sum_value)))
                self.table_widget.setItem(i, self.table_widget.columnCount() - 2, QTableWidgetItem(str(tolerance)))


        else:
            QMessageBox.warning(self, '错误', "没有可计算的数据")
            # 可以在这里显示结果，例如通过弹窗或更新表格


    def showAbout(self):
        QMessageBox.about(self, "关于", "这是一个程序")
    # 使用函数处理数据



if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        main_window = MainWindow()
        sys.exit(app.exec_())
    except Exception as e:
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle("错误")
        msg_box.setText(f"程序启动失败：{str(e)}")
        msg_box.exec()
